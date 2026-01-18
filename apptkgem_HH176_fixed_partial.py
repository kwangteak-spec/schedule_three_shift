import tkinter as tk
from tkinter import ttk, messagebox
import calendar
import json
import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment 
    
import subprocess
from holidays import is_holiday


# --- 근무시간 상수 (원하시면 값 변경 가능) ---
DAY_HOURS = 8.0                # 주간(오전/오후) 근무 시간
NIGHT_BASE_HOURS = 8.0         # 야간 근무의 기본 근무시간 (예: 8시간)
NIGHT_ALLOWANCE_HOURS = 1.5    # 야간 수당(추가)
NIGHT_TOTAL_HOURS = NIGHT_BASE_HOURS + NIGHT_ALLOWANCE_HOURS

# --- 설정값 ---
WORK_STATS_FILE = "work_stats_data.json" # <--- 내부 근무 통계 및 달력 상태 저장용 파일명

# --- 1. 데이터 불러오기 및 저장 함수 정의 ---
def load_data():
    """
    각종 데이터를 파일에서 불러오는 함수
    """
    staff_list = []
    night_support_assignments = {}
    additional_off_days = {}
    additional_work_days = {}
    work_stats = {}

    # work_stats (work_stats_data.json)
    # 현재 달력의 년도와 월을 먼저 설정하기 위해 work_stats_data.json을 먼저 불러옴
    work_result_file = WORK_STATS_FILE # 통일된 변수 사용
    if os.path.exists(work_result_file):
        with open(work_result_file, "r", encoding="utf-8") as f:
            try:
                work_stats = json.load(f)
            except json.JSONDecodeError:
                # 오류 발생 시 빈 딕셔너리로 초기화하고 경고 메시지를 출력
                work_stats = {}
                print(f"경고: '{WORK_STATS_FILE}' 파일의 형식이 올바르지 않습니다. 새로운 데이터로 초기화합니다.")

    # work_stats에 년도와 월이 없으면 현재 날짜로 설정
    now = datetime.now()
    current_year = work_stats.get('year', now.year)
    current_month = work_stats.get('month', now.month)

    # staff_list (name00.txt)
    name_file = "name00.txt"
    if os.path.exists(name_file):
        with open(name_file, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip() and not line.startswith("기본사용자"):
                    try:
                        parts = line.strip().split(',')
                        name = parts[0]
                        shift = parts[1]
                        off_days = []
                        if len(parts) > 2 and parts[2]:
                            off_days = [int(d) for d in parts[2].split('|')]
                        staff_list.append({'name': name, 'shift': shift, 'off_days': off_days})
                    except (ValueError, IndexError):
                        print(f"경고: name00.txt 파일의 형식이 올바르지 않습니다: {line.strip()}")
    else:
        print("경고: 'name00.txt' 파일을 찾을 수 없습니다. 새로운 기본 데이터로 실행합니다.")

    # work_stats에 'names' 키가 없으면 새로 생성
    if 'names' not in work_stats or not work_stats['names']:
        work_stats = {
            "year": current_year,
            "month": current_month,
            "names": {s['name']: {"work_days": 0, "work_hours": 0} for s in staff_list}
        }

    # night_support_assignments (afternoon_night_support.txt)
    support_file = "afternoon_night_support.txt"
    if os.path.exists(support_file):
        with open(support_file, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    try:
                        date_str, name = line.strip().split(',')
                        support_date = datetime.strptime(date_str.strip(), "%Y-%m-%d")

                        # 년도와 월이 일치하는 데이터만 불러옴
                        if support_date.year == current_year and support_date.month == current_month:
                            if name not in night_support_assignments:
                                night_support_assignments[name] = []
                            night_support_assignments[name].append(support_date.day)
                    except (ValueError, IndexError):
                        print(f"경고: afternoon_night_support.txt 파일의 형식이 올바르지 않습니다: {line.strip()}")

    # additional_off_days (addday.json)
    add_day_file = "addday.json"
    if os.path.exists(add_day_file):
        with open(add_day_file, "r", encoding="utf-8") as f:
            try:
                additional_off_days = json.load(f)
            except json.JSONDecodeError:
                print("경고: 'addday.json' 파일의 형식이 올바르지 않습니다.")

    # additional_work_days (deadd.json)
    work_day_file = "deadd.json"
    if os.path.exists(work_day_file):
        with open(work_day_file, "r", encoding="utf-8") as f:
            try:
                additional_work_days = json.load(f) # JSON 파일 로드
            except json.JSONDecodeError:
                print("경고: 'deadd.json' 파일의 형식이 올바르지 않습니다.")
            except Exception as e:
                # 파일은 존재하지만 내용이 비어있거나 다른 예상치 못한 오류 발생 시
                print(f"경고: 'deadd.json' 파일을 불러오는 중 오류 발생: {e}")

    return staff_list, night_support_assignments, additional_off_days, additional_work_days, work_stats, current_year, current_month

def load_support_rotation_map(filename="support_rotation_map.json"):
    """
    지원 근무자의 근무 패턴을 JSON 파일에서 불러오는 함수
    """
    if os.path.exists(filename):
        with open(filename, "r", encoding="utf-8") as f:
            try:
                # JSON 키는 문자열이므로, 정수형으로 변환하여 딕셔너리를 생성
                return {int(k): v for k, v in json.load(f).items()}
            except (json.JSONDecodeError, ValueError):
                print(f"경고: '{filename}' 파일의 형식이 올바르지 않습니다. 기본 패턴을 사용합니다.")
    else:
        print(f"경고: '{filename}' 파일을 찾을 수 없습니다. 기본 패턴을 사용합니다.")
    # 파일이 없거나 오류가 발생하면 기본값 반환
    return {
        0: '오후',
        1: '오후',
        2: '오후',
        3: '휴무',
        4: '휴무',
        5: '오전',
        6: '오전'
    }

def save_data(staff_list):
    """
    직원 데이터를 name00.txt에 저장하는 함수
    """
    name_file = "name00.txt"
    with open(name_file, "w", encoding="utf-8") as f:
        f.write("기본사용자,선택하세요,\n")
        for staff in staff_list:
            off_days_str = '|'.join(map(str, staff['off_days']))
            f.write(f"{staff['name']},{staff['shift']},{off_days_str}\n")
    print(f"근무자 데이터가 '{name_file}'에 저장되었습니다.")

def save_night_support_data(night_support_assignments, year, month):
    """
    야간 지원 데이터를 afternoon_night_support.txt에 저장하는 함수
    """
    support_file = "afternoon_night_support.txt"
    existing_data = []
    if os.path.exists(support_file):
        with open(support_file, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    date_str, _ = line.strip().split(',')
                    support_date = datetime.strptime(date_str.strip(), "%Y-%m-%d")

                    if support_date.year != year or support_date.month != month:
                        existing_data.append(line.strip())
                except (ValueError, IndexError):
                    pass

    for name, days in night_support_assignments.items():
        for day in days:
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            existing_data.append(f"{date_str},{name}")

    with open(support_file, "w", encoding="utf-8") as f:
        for line in existing_data:
            f.write(f"{line}\n")
    print(f"야간 지원 데이터가 '{support_file}'에 저장되었습니다.")

def save_additional_off_days(additional_off_days):
    """
    추가 휴무일 데이터를 addday.json에 저장하는 함수
    """
    add_day_file = "addday.json"
    with open(add_day_file, "w", encoding="utf-8") as f:
        json.dump(additional_off_days, f, ensure_ascii=False, indent=4)
    print(f"추가 휴무일 데이터가 '{add_day_file}'에 저장되었습니다.")

def save_additional_work_days(additional_work_days):
    """
    추가 근무일 데이터를 deadd.json에 저장하는 함수
    """
    work_day_file = "deadd.json"
    with open(work_day_file, "w", encoding="utf-8") as f:
        json.dump(additional_work_days, f, ensure_ascii=False, indent=4)
    print(f"추가 근무일 데이터가 '{work_day_file}'에 저장되었습니다.")

def save_work_stats(work_stats):
    """
    근무 통계 데이터를 work_stats_data.json에 저장하는 함수
    """
    work_result_file = WORK_STATS_FILE # 통일된 변수 사용
    with open(work_result_file, "w", encoding="utf-8") as f:
        json.dump(work_stats, f, ensure_ascii=False, indent=4)
    print(f"근무 통계 데이터가 '{work_result_file}'에 저장되었습니다.")


# --- 2. GUI 클래스 정의 ---

class WorkScheduleApp(tk.Tk):
    def __init__(self, staff_list, night_support_assignments, additional_off_days, additional_work_days, work_stats, current_year, current_month):
        super().__init__()
        self.title("근무자 일정 관리 프로그램")
        self.geometry("1024x768")
        self.staff_list = staff_list
        self.night_support_assignments = night_support_assignments
        self.additional_off_days = additional_off_days
        self.additional_work_days = additional_work_days
        self.work_stats = work_stats

        self.current_year = current_year
        self.current_month = current_month
        
        # 년/월 콤보박스 변수 초기화 (create_widgets에서 사용)
        self.current_year_var = tk.IntVar(value=self.current_year)
        self.current_month_var = tk.IntVar(value=self.current_month)


        # staff_list의 인원 수에 따라 초기 모드 자동 설정
        num_staff = len(self.staff_list)
        if num_staff <= 5:
            self.current_mode = "비상 모드 (5인 이하)"
        else:
            self.current_mode = "기본 모드 (5인 이상)"

        self.rotating_staff_names = [s['name'] for s in self.staff_list if s['shift'] == '지원']
        self.rotating_staff_names.sort()

        # '지원' 근무자의 새로운 근무 패턴을 파일에서 불러옴
        self.support_rotation_map = load_support_rotation_map()

        night_staff_names = sorted([s['name'] for s in self.staff_list if s['shift'].startswith('야간')])
        self.fixed_night_schedule = []
        if len(night_staff_names) >= 2:
            night_person_1 = night_staff_names[0]
            night_person_2 = night_staff_names[1]
            for i in range(31):
                if (i // 2) % 2 == 0:
                    self.fixed_night_schedule.append(night_person_1)
                else:
                    self.fixed_night_schedule.append(night_person_2)

        self.night_shift_index = 0

        self.create_widgets()

    # --- 새로 추가된 년/월 변경 함수 ---
    def set_new_calendar_date(self, event=None):
        """
        년도와 월을 변경하고, work_stats_data.json에 저장 후 달력을 업데이트합니다.
        """
        try:
            new_year = self.current_year_var.get()
            new_month = self.current_month_var.get()

            # 현재 값과 다를 경우에만 처리
            if new_year != self.current_year or new_month != self.current_month:
                self.current_year = new_year
                self.current_month = new_month

                # 1. work_stats_data.json에 변경된 년/월 저장
                self.work_stats['year'] = self.current_year
                self.work_stats['month'] = self.current_month
                save_work_stats(self.work_stats)

                # 2. 월이 변경되었으므로, 새 월에 해당하는 야간 지원 및 기타 데이터를 파일에서 다시 불러와야 함
                self.reload_data_for_new_month() # <--- 이 함수에서 추가 파일 초기화 작업 수행
                
                # 3. 변경된 날짜로 UI 업데이트
                self.update_calendar()
                self.update_off_days_section() # 선택 직원 통계 업데이트
                
                messagebox.showinfo("날짜 변경", f"{self.current_year}년 {self.current_month}월 근무표로 화면이 갱신되었습니다.")

        except Exception as e:
            messagebox.showerror("날짜 변경 오류", f"년도 또는 월 변경 중 오류 발생: {e}")
            
    def reload_data_for_new_month(self):
        """
        년/월이 변경되었을 때 해당 월에 대한 데이터를 새로 불러오고,
        월별로 초기화되어야 하는 파일들(addday.json, deadd.json)을 초기화합니다.
        """
        # 임시로 현재 년/월을 저장
        temp_year = self.current_year
        temp_month = self.current_month
        
        # night_support_assignments 업데이트 (load_data 로직의 일부) - 월별 필터링
        self.night_support_assignments = {}
        support_file = "afternoon_night_support.txt"
        if os.path.exists(support_file):
            with open(support_file, "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip():
                        try:
                            date_str, name = line.strip().split(',')
                            support_date = datetime.strptime(date_str.strip(), "%Y-%m-%d")

                            if support_date.year == temp_year and support_date.month == temp_month:
                                if name not in self.night_support_assignments:
                                    self.night_support_assignments[name] = []
                                self.night_support_assignments[name].append(support_date.day)
                        except:
                            pass
        
        # --- 추가 휴무일 (addday.json) 초기화 및 저장 ---
        self.additional_off_days = {}
        save_additional_off_days(self.additional_off_days)

        # --- 추가 근무일 (deadd.json) 초기화 및 저장 ---
        self.additional_work_days = {}
        save_additional_work_days(self.additional_work_days)
        
        # 지원 근무 패턴 재로드
        self.support_rotation_map = load_support_rotation_map()
    
    def export_to_excel(self):
        # 1. 엑셀 파일 이름 및 시트 이름 설정
        excel_file_path = "엑셀근무표.xlsx"
        sheet_name = f"{self.current_year}년_{self.current_month}월"

        # 2. 날짜별 근무자 정보 추출 (데이터 수집)
        import calendar
        cal = calendar.Calendar()
        
        daily_data = {
            '일자': [],
            '요일': [],
            '오전근무': [],
            '오후근무': [],
            '야간근무': []
        }
        
        day_names = ["월", "화", "수", "목", "금", "토", "일"]
        self.night_shift_index = 0 
        
        for day in cal.itermonthdates(self.current_year, self.current_month):
            if day.month == self.current_month:
                staff_on_duty = self.get_staff_for_day(day)
                
                daily_data['일자'].append(day.day)
                daily_data['요일'].append(day_names[day.weekday()])
                daily_data['오전근무'].append('\n'.join(staff_on_duty.get('오전', [])))
                daily_data['오후근무'].append('\n'.join(staff_on_duty.get('오후', [])))
                daily_data['야간근무'].append('\n'.join(staff_on_duty.get('야간', [])))

        # 3. 데이터프레임 구성 및 변환 (행/열 반전)
        df = pd.DataFrame(daily_data)
        df_transposed = df.set_index(['일자', '요일']).T
        
        # Openpyxl에 쓰기 위해 DataFrame을 리스트 오브 리스트로 변환
        data_to_write = [
            # 년/월 정보를 포함하는 헤더 행
            ["CU편의점", "", f"{self.current_year}년 {self.current_month}월 근무표"], 
            # 빈 행
            [], 
            # 컬럼 헤더 (구분, 날짜 1, 날짜 2, ...)
            ['구분'] + list(df_transposed.columns.get_level_values('일자').astype(str)),
            # 근무조 행
            ['요일'] + list(df_transposed.columns.get_level_values('요일')),
        ]
        
        # 근무조 데이터 추가
        for shift, row in df_transposed.iterrows():
            data_to_write.append([shift] + row.tolist())

        # 4. 엑셀 파일 저장 및 업데이트 (Openpyxl 단독 사용)
        try:
            # 4-1. 기존 파일 로드 또는 새 파일 생성
            if os.path.exists(excel_file_path):
                # 기존 파일을 로드할 때 VBA 유지 및 읽기 전용 해제
                workbook = openpyxl.load_workbook(excel_file_path, read_only=False, keep_vba=True)
                
                # 기존 시트 삭제 (업데이트)
                if sheet_name in workbook.sheetnames:
                    workbook.remove(workbook[sheet_name])
            else:
                # 파일이 없으면 새로운 워크북 생성
                workbook = openpyxl.Workbook()
                # 기본 시트 삭제
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
            
            # 4-2. 새로운 시트 생성 및 활성화
            worksheet = workbook.create_sheet(sheet_name) 

            # 4-3. 데이터 쓰기 (셀 단위)
            start_row = 1
            for row_data in data_to_write:
                # 첫 번째 행은 CU편의점 및 년월 정보
                if start_row == 1:
                    worksheet.cell(row=1, column=1, value=row_data[0]) # CU편의점
                    worksheet.cell(row=1, column=3, value=row_data[2]) # 년월 근무표
                
                # 세 번째 행부터 실제 데이터 헤더 시작
                elif start_row >= 3:
                    # 셀 인덱스는 1부터 시작
                    for col_idx, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=start_row, column=col_idx, value=value)
                        
                        # 실제 근무 데이터가 시작되는 5번째 행부터 서식 적용
                        if start_row >= 5:
                            from openpyxl.styles import Alignment 
                            cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
                        
                start_row += 1

            # 4-4. 서식 적용
            from openpyxl.utils import get_column_letter 
            
            # '구분' (근무조) 셀 너비 조정
            worksheet.column_dimensions['A'].width = 10
            
            # '일자', '요일' 셀 스타일 조정 (셀 폭에 따라)
            # 데이터 열 (B열부터 끝까지)
            for col_idx in range(2, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 10

            # 4-5. 최종적으로 워크북을 저장
            # Openpyxl 객체가 모든 파일 핸들링을 전담하므로 충돌 가능성 최소화
            workbook.save(excel_file_path)
            
            messagebox.showinfo("엑셀 출력 완료", f"근무표가 '{excel_file_path}' 파일의 '{sheet_name}' 시트로 저장되었습니다.")
            subprocess.Popen(['start', excel_file_path], shell=True)
            
        except Exception as e:
            # 오류 발생 시 사용자에게 팝업으로 명확히 알림
            messagebox.showerror("오류", f"엑셀 파일을 저장하는 중 오류가 발생했습니다: {e}")
    
    
    
    def show_legal_hours(self):
        """
        [MODIFIED] 'work_result.json' 파일의 내용을 메시지 상자에 표시합니다.
        월별 법정 근무 통계를 보여줍니다.
        """
        # 사용자 요청에 따라 'work_result.json' 파일을 명시적으로 사용합니다.
        work_result_file = "work_result.json" 

        # 1. 파일이 존재하는지 확인
        if not os.path.exists(work_result_file):
            messagebox.showerror("오류", f"'{work_result_file}' 파일을 찾을 수 없습니다. 법정근무시간 계산이 완료되었는지 확인해주세요.")
            return

        try:
            # 2. JSON 파일 내용 불러오기
            with open(work_result_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            # 3. 필요한 키들이 모두 있는지 확인 (선택 사항)
            # 4. 각 통계 값 가져오기 (키가 없을 경우 기본값 0 또는 빈 문자열 사용)
            year = data.get('year', 'N/A')
            month = data.get('month', 'N/A')
            workdays = data.get('workdays', 0)
            total_hours = data.get('total_hours', 0)
            night_support_days = data.get('night_support_days', 0)
            night_leave_hours = data.get('night_leave_hours', 0)

          
            # 5. 출력할 메시지 형식 만들기
            message = (
                f"--- {year}년 {month}월 근무 통계 ---\n\n"
                f"법정근무일: {workdays}일\n"
                f"법정근무시간: {total_hours}시간\n"
                f"오후조 야간지원일: {night_support_days}일\n"
                f"야간조 월차시간: {night_leave_hours}시간\n"
            )
            

            # 6. 메시지 팝업창으로 띄우기
            messagebox.showinfo("법정근무시간", message)

        except json.JSONDecodeError:
            messagebox.showerror("오류", "파일 형식이 올바르지 않습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"파일을 읽는 중 오류가 발생했습니다: {e}")
    
    
    
    def start_new_work(self):
        """
        새로운 월의 근무표 작업을 시작하는 함수
        """
        response = messagebox.askyesno("새로운 작업 시작", "기존의 모든 근무표 정보가 초기화됩니다. 계속하시겠습니까?")
        if response:
            self.start_new_work_auto()
    
    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        control_frame = ttk.Frame(main_frame, padding=10)
        control_frame.pack(side=tk.LEFT, fill=tk.Y)

        ttk.Label(control_frame, text="근무표 관리", font=("Arial", 14, "bold")).pack(pady=5)
        
        # --- [NEW] 년도/월 선택 콤보박스 추가 ---
        date_frame = ttk.Frame(control_frame)
        date_frame.pack(pady=10)

        # Year selection
        ttk.Label(date_frame, text="년도").pack(side=tk.LEFT, padx=(0, 5))
        self.year_combo = ttk.Combobox(date_frame, textvariable=self.current_year_var, state="readonly", width=5, justify='center')
        self.year_combo['values'] = list(range(self.current_year - 2, self.current_year + 3)) # 현재 년도 기준 5년치
        self.year_combo.bind("<<ComboboxSelected>>", self.set_new_calendar_date)
        self.year_combo.pack(side=tk.LEFT, padx=(0, 10))

        # Month selection
        ttk.Label(date_frame, text="월").pack(side=tk.LEFT, padx=(0, 5))
        self.month_combo = ttk.Combobox(date_frame, textvariable=self.current_month_var, state="readonly", width=3, justify='center')
        self.month_combo['values'] = list(range(1, 13))
        self.month_combo.bind("<<ComboboxSelected>>", self.set_new_calendar_date)
        self.month_combo.pack(side=tk.LEFT)
        # ----------------------------------------


        ttk.Label(control_frame, text="근무 모드 선택").pack(pady=(5, 0))
        self.mode_combo = ttk.Combobox(control_frame, values=["기본 모드 (5인 이상)", "비상 모드 (5인 이하)"])
        self.mode_combo.set(self.current_mode)
        self.mode_combo.bind("<<ComboboxSelected>>", self.set_mode)
        self.mode_combo.pack()

        ttk.Label(control_frame, text="근무조 선택").pack(pady=(5, 0))
        self.shift_combo = ttk.Combobox(control_frame, values=["모든 근무조", "오전조", "오후조", "야간조"])
        self.shift_combo.set("모든 근무조")
        self.shift_combo.bind("<<ComboboxSelected>>", self.update_calendar)
        self.shift_combo.pack()

        ttk.Label(control_frame, text="직원 선택").pack(pady=(10, 0))
        self.staff_combo = ttk.Combobox(control_frame, values=[s['name'] for s in self.staff_list])
        self.staff_combo.bind("<<ComboboxSelected>>", self.update_off_days_section)
        self.staff_combo.pack()

        self.add_off_days_frame = ttk.LabelFrame(control_frame, text="추가 휴무일 지정 (날짜)", padding=5)
        self.add_off_days_frame.pack(pady=10, fill=tk.X)

        ttk.Label(self.add_off_days_frame, text="날짜 (쉼표 구분):").pack(anchor="w")
        self.add_off_entry = ttk.Entry(self.add_off_days_frame)
        self.add_off_entry.pack(fill=tk.X, pady=5)

        self.add_off_button = ttk.Button(self.add_off_days_frame, text="추가 휴무일 적용", command=self.apply_additional_off_days)
        self.add_off_button.pack(pady=5, fill=tk.X)

        self.add_work_days_frame = ttk.LabelFrame(control_frame, text="추가 근무일 지정 (날짜)", padding=5)
        self.add_work_days_frame.pack(pady=10, fill=tk.X)

        ttk.Label(self.add_work_days_frame, text="날짜 (쉼표 구분):").pack(anchor="w")
        self.add_work_entry = ttk.Entry(self.add_work_days_frame)
        self.add_work_entry.pack(fill=tk.X, pady=5)

        self.add_work_button = ttk.Button(self.add_work_days_frame, text="추가 근무일 적용", command=self.apply_additional_work_days)
        self.add_work_button.pack(pady=5, fill=tk.X)

        self.night_support_frame = ttk.LabelFrame(control_frame, text="야간 지원 지정 (오후조)", padding=5)
        # 이 프레임은 초기에는 숨깁니다.
        self.night_support_frame.pack_forget()

        ttk.Label(self.night_support_frame, text="날짜 (쉼표 구분):").pack(anchor="w")
        self.night_support_entry = ttk.Entry(self.night_support_frame)
        self.night_support_entry.pack(fill=tk.X, pady=5)

        self.night_support_button = ttk.Button(self.night_support_frame, text="야간 지원 적용", command=self.apply_night_support)
        self.night_support_button.pack(pady=5, fill=tk.X)

        stats_frame = ttk.LabelFrame(control_frame, text="근무시간 통계 (선택 직원)", padding=5)
        stats_frame.pack(pady=10, fill=tk.X)
        self.stats_label = ttk.Label(stats_frame, text="")
        self.stats_label.pack()

        ttk.Separator(control_frame, orient='horizontal').pack(fill='x', pady=5)

        ttk.Button(control_frame, text="새로운 작업 시작", command=self.start_new_work).pack(pady=5, fill=tk.X)

        self.calendar_frame = ttk.Frame(main_frame)
        self.calendar_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.update_calendar()
        self.update_total_stats()

    
    def calculate_legal_work_stats(self):
        """
        현재 월의 법정근무일수와 법정근무시간을 계산하여 self.work_stats에 저장합니다.
        (주 5일, 주 40시간/일 8시간 기준)
        """
        cal = calendar.Calendar()
        month_work_days = 0
        
        # 월~금요일(0~4)만 계산에 포함
        for day in cal.itermonthdates(self.current_year, self.current_month):
            if day.month == self.current_month and day.weekday() < 5:
                # 공휴일인지 확인
                from datetime import date
                if not is_holiday(date(day.year, day.month, day.day)):
                    month_work_days += 1
        
        # 법정근무일과 법정근무시간 계산
        self.work_stats['workdays'] = month_work_days 
        self.work_stats['total_hours'] = month_work_days * 8
        # night_support_days와 night_leave_hours는 이 함수에서 계산하지 않음 (기존 로직 유지)
    
    def set_mode(self, event=None):
        self.current_mode = self.mode_combo.get()
        self.update_night_support_frame_visibility()
        self.update_calendar()
        self.update_total_stats()

    def update_night_support_frame_visibility(self, event=None):
        selected_name = self.staff_combo.get()
        selected_staff = next((s for s in self.staff_list if s['name'] == selected_name), None)

        if self.current_mode == "기본 모드 (5인 이상)" and selected_staff and '오후' in selected_staff['shift']:
            self.night_support_frame.pack(pady=10, fill=tk.X)
        else:
            self.night_support_frame.pack_forget()

    def update_off_days_section(self, event=None):
        selected_name = self.staff_combo.get()
        if not selected_name:
            self.stats_label.config(text="직원을 선택하면 통계가 표시됩니다.")
            return

        selected_staff = next((s for s in self.staff_list if s['name'] == selected_name), None)

        self.add_off_entry.delete(0, tk.END)
        if selected_name in self.additional_off_days:
            dates_str = ', '.join(map(str, self.additional_off_days[selected_name]))
            self.add_off_entry.insert(0, dates_str)

        self.add_work_entry.delete(0, tk.END)
        if selected_name in self.additional_work_days:
            dates_str = ', '.join(map(str, self.additional_work_days[selected_name]))
            self.add_work_entry.insert(0, dates_str)

        self.night_support_entry.delete(0, tk.END)
        if selected_name in self.night_support_assignments:
            dates_str = ', '.join(map(str, self.night_support_assignments[selected_name]))
            self.night_support_entry.insert(0, dates_str)

        self.update_night_support_frame_visibility()
        self.update_stats()

    def apply_additional_off_days(self):
        selected_name = self.staff_combo.get()
        if not selected_name:
            messagebox.showerror("오류", "먼저 직원을 선택해주세요.")
            return

        dates_str = self.add_off_entry.get()
        try:
            dates = [int(d.strip()) for d in dates_str.split(',') if d.strip()]
            self.additional_off_days[selected_name] = dates
            save_additional_off_days(self.additional_off_days)
            messagebox.showinfo("저장 완료", "추가 휴무일이 성공적으로 저장되었습니다.")
            self.update_calendar()
            self.update_stats()
        except ValueError:
            messagebox.showerror("오류", "날짜는 쉼표로 구분된 숫자로 입력해주세요.")

    def apply_additional_work_days(self):
        selected_name = self.staff_combo.get()
        if not selected_name:
            messagebox.showerror("오류", "먼저 직원을 선택해주세요.")
            return

        dates_str = self.add_work_entry.get()
        try:
            dates = [int(d.strip()) for d in dates_str.split(',') if d.strip()]

            # 선택된 직원의 근무조 정보 가져오기
            selected_staff = next((s for s in self.staff_list if s['name'] == selected_name), None)

            # --- 야간조 추가 근무일 특별 처리 로직 ---
            if selected_staff and selected_staff['shift'].startswith('야간'):
                # 야간조1, 야간조2 이름 찾기
                night_staff_names = sorted([s['name'] for s in self.staff_list if s['shift'].startswith('야간')])
                if len(night_staff_names) >= 2:
                    night_person_1, night_person_2 = night_staff_names[0], night_staff_names[1]

                    # 근무일수 계산
                    work_days_1 = 0
                    work_days_2 = 0
                    cal = calendar.Calendar()
                    for day in cal.itermonthdates(self.current_year, self.current_month):
                        if day.month == self.current_month:
                            if self.fixed_night_schedule[day.day - 1] == night_person_1:
                                work_days_1 += 1
                            elif self.fixed_night_schedule[day.day - 1] == night_person_2:
                                work_days_2 += 1

                    # 근무일수 차이가 2일 이상인지 확인
                    if abs(work_days_1 - work_days_2) >= 2:
                        # 추가 근무일을 야간 지원으로 처리
                        self.night_support_assignments[selected_name] = dates
                        save_night_support_data(self.night_support_assignments, self.current_year, self.current_month)
                        messagebox.showinfo("저장 완료", f"야간조 근무일수 불균형으로 인해\n{selected_name}님의 추가 근무일이 야간 지원으로 저장되었습니다.")
                        self.update_calendar()
                        self.update_stats()
                        return

            # 일반적인 추가 근무일 로직
            self.additional_work_days[selected_name] = dates
            save_additional_work_days(self.additional_work_days)
            messagebox.showinfo("저장 완료", "추가 근무일이 성공적으로 저장되었습니다.")
            self.update_calendar()
            self.update_stats()

        except ValueError:
            messagebox.showerror("오류", "날짜는 쉼표로 구분된 숫자로 입력해주세요.")

    def apply_night_support(self):
        selected_name = self.staff_combo.get()
        if not selected_name:
            messagebox.showerror("오류", "먼저 직원을 선택해주세요.")
            return

        dates_str = self.night_support_entry.get()
        try:
            dates = [int(d.strip()) for d in dates_str.split(',') if d.strip()]
            self.night_support_assignments[selected_name] = dates
            save_night_support_data(self.night_support_assignments, self.current_year, self.current_month)
            messagebox.showinfo("저장 완료", "야간 지원 날짜가 성공적으로 저장되었습니다.")
            self.update_calendar()
            self.update_stats()
        except ValueError:
            messagebox.showerror("오류", "날짜는 쉼표로 구분된 숫자로 입력해주세요.")

    # --- START MODIFIED update_stats ---
    def update_stats(self):
        selected_name = self.staff_combo.get()
        if not selected_name:
            self.stats_label.config(text="직원을 선택하면 통계가 표시됩니다.")
            return

        total_work_days = 0
        total_work_hours = 0

        cal = calendar.Calendar()
        self.night_shift_index = 0

        for day in cal.itermonthdates(self.current_year, self.current_month):
            if day.month != self.current_month:
                continue

            staff_on_duty = self.get_staff_for_day(day)

            # 1. 야간조 (Night Shift, 12시간) 근무 체크:
            # selected_name이 그날의 야간 근무자(고정 or 지원)에 포함되는 경우
            # get_staff_for_day가 야간 지원 근무자도 '야간' 쉬프트에 배치하므로 이 체크가 누락을 방지함.
            if selected_name in staff_on_duty.get("야간", []):
                total_work_days += 1
                total_work_hours += 12
            
            # 2. 주간조 (Day Shift, 8시간) 근무 체크:
            # selected_name이 오전 또는 오후 근무에 포함되는 경우
            elif selected_name in staff_on_duty.get("오전", []) or selected_name in staff_on_duty.get("오후", []):
                total_work_days += 1
                total_work_hours += 8
            
            # Note: 야간 지원을 수행하는 경우 (12시간으로 카운트됨)는 elif로 진입하지 않음.

        self.stats_label.config(text=
            f"총 근무일: {total_work_days}일\n"
            f"총 근무시간: {total_work_hours}시간"
        )

        if selected_name in self.work_stats['names']:
            self.work_stats['names'][selected_name]['work_days'] = total_work_days
            self.work_stats['names'][selected_name]['work_hours'] = total_work_hours
    # --- END MODIFIED update_stats ---


    # --- START MODIFIED update_total_stats ---
    def update_total_stats(self):
        total_work_days_all = {}
        total_work_hours_all = {}

        cal = calendar.Calendar()
        self.night_shift_index = 0

        for staff in self.staff_list:
            name = staff['name']
            total_work_days = 0
            total_work_hours = 0

            for day in cal.itermonthdates(self.current_year, self.current_month):
                if day.month != self.current_month:
                    continue

                staff_on_duty = self.get_staff_for_day(day)

                # 1. 야간조 (Night Shift, 12시간) 근무 체크:
                # name이 그날의 야간 근무자(고정 or 지원)에 포함되는 경우
                if name in staff_on_duty.get("야간", []):
                    total_work_days += 1
                    total_work_hours += 12
                
                # 2. 주간조 (Day Shift, 8시간) 근무 체크:
                # name이 오전 또는 오후 근무에 포함되는 경우
                elif name in staff_on_duty.get("오전", []) or name in staff_on_duty.get("오후", []):
                    total_work_days += 1
                    total_work_hours += 8
                
            total_work_days_all[name] = total_work_days
            total_work_hours_all[name] = total_work_hours

            if name in self.work_stats['names']:
                self.work_stats['names'][name]['work_days'] = total_work_days
                self.work_stats['names'][name]['work_hours'] = total_work_hours
        
        # [MODIFIED] 법정근무일/시간 계산 및 갱신 로직 추가
        self.calculate_legal_work_stats()

        save_work_stats(self.work_stats)

        stats_text = ""
        for name in sorted(total_work_days_all.keys()):
            stats_text += f"{name}: {total_work_days_all[name]}일, {total_work_hours_all[name]:.1f}시간\n"

        messagebox.showinfo("전체 근무표 통계", stats_text)
    # --- END MODIFIED update_total_stats ---

    # --- [MODIFIED] 외부 스크립트 실행 후 새로고침 로직 추가 ---
    def run_other_script(self):
        """
        'inittotallastsch_ws46.py' 스크립트를 실행하고 UI를 새로고침합니다.
        """
        try:
            script_path = 'inittotallastsch_ws46.py'
            messagebox.showinfo("실행", f"'{script_path}' 파일을 실행합니다. 완료될 때까지 기다려주세요. 이 작업은 근무 패턴을 업데이트합니다.")
            
            # subprocess.run을 사용하여 외부 스크립트가 완료될 때까지 대기 (Blocking)
            result = subprocess.run(
                ['python', script_path], 
                check=True, 
                capture_output=True, 
                text=True, 
                encoding='utf-8'
            )
            
            # 실행 성공 메시지
            messagebox.showinfo("실행 완료", f"'{script_path}' 파일 실행이 완료되었습니다. 근무표를 새로고침합니다.")
            
            # 데이터 다시 불러오기 및 UI 업데이트
            self.reload_data_and_ui()

        except FileNotFoundError:
            messagebox.showerror("오류", f"'{script_path}' 파일을 찾을 수 없습니다. 경로를 확인해주세요.")
        except subprocess.CalledProcessError as e:
            # 외부 스크립트에서 오류 발생 시 출력
            error_output = e.stderr if e.stderr else e.stdout
            messagebox.showerror("실행 오류", f"파일 실행 중 오류가 발생했습니다. 자세한 내용:\n{error_output[-500:]}")
        except Exception as e:
            messagebox.showerror("오류", f"파일 실행 및 업데이트 중 예기치 않은 오류가 발생했습니다: {e}")
    # -------------------------------------------------------------

    def reload_data_and_ui(self):
        """
        데이터를 다시 불러오고 UI를 새로고침하는 함수
        """
        # 기존 데이터를 모두 초기화
        self.staff_list.clear()
        self.night_support_assignments.clear()
        self.additional_off_days.clear()
        self.additional_work_days.clear()
        self.work_stats.clear()

        # 파일에서 최신 데이터 다시 불러오기
        (staff_list, self.night_support_assignments, 
         self.additional_off_days, self.additional_work_days, 
         self.work_stats, self.current_year, self.current_month) = load_data()
         
        # staff_list 업데이트 (clear 후 새로 할당)
        self.staff_list.extend(staff_list)
        
        # 년/월 변수 업데이트
        self.current_year_var.set(self.current_year)
        self.current_month_var.set(self.current_month)

        # UI 업데이트
        self.update_calendar()
        self.update_off_days_section()
        self.update_total_stats()

    def update_calendar(self, event=None):
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()

        title_frame = ttk.Frame(self.calendar_frame)
        title_frame.pack(fill=tk.X, pady=5)

        ttk.Label(title_frame, text=f"{self.current_year}년 {self.current_month}월", font=("Arial", 14, "bold")).pack(side=tk.LEFT, padx=10)

        # "근무 패턴 관리 실행" 버튼 추가
        run_pattern_button = ttk.Button(title_frame, text="근무 패턴 관리 실행", command=self.run_other_script)
        run_pattern_button.pack(side=tk.RIGHT, padx=5)

        # "법정근무시간" 버튼 추가
        legal_hours_button = ttk.Button(title_frame, text="법정근무시간", command=self.show_legal_hours)
        legal_hours_button.pack(side=tk.RIGHT, padx=5)

        # "전체 근무표 통계" 버튼 추가
        total_stats_button = ttk.Button(title_frame, text="전체 근무표 통계", command=self.update_total_stats)
        total_stats_button.pack(side=tk.RIGHT, padx=5)

        excel_export_button = ttk.Button(title_frame, text="근무표 Excel로 내보내기", command=self.export_to_excel)
        excel_export_button.pack(side=tk.RIGHT, padx=5)

        cal_grid = ttk.Frame(self.calendar_frame)
        cal_grid.pack(fill=tk.BOTH, expand=True)

        calendar.setfirstweekday(calendar.MONDAY)

        days_of_week = ["월", "화", "수", "목", "금", "토", "일"]
        for i, day in enumerate(days_of_week):
            ttk.Label(cal_grid, text=day, font=("Arial", 9, "bold"), width=12, anchor="center").grid(row=0, column=i, sticky="nsew")

        cal = calendar.Calendar()
        dates_in_month = cal.itermonthdates(self.current_year, self.current_month)

        row, col = 1, 0
        self.night_shift_index = 0

        for day in dates_in_month:
            if day.month == self.current_month:
                date_text = str(day.day)
                is_current_month = True
            else:
                date_text = ""
                is_current_month = False

            cell_frame = ttk.Frame(cal_grid, borderwidth=1, relief="solid")
            cell_frame.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)

            date_label = ttk.Label(cell_frame, text=date_text, font=("Arial", 10, "bold"))
            date_label.pack(side=tk.TOP, anchor="ne", padx=2, pady=2)

            if is_current_month:
                staff_on_duty = self.get_staff_for_day(day)

                filter_type = self.shift_combo.get()

                ordered_shifts = []
                if "오전" in staff_on_duty: ordered_shifts.append("오전")
                if "오후" in staff_on_duty: ordered_shifts.append("오후")
                if "야간" in staff_on_duty: ordered_shifts.append("야간")

                for shift_name in ordered_shifts:
                    names = staff_on_duty[shift_name]

                    if filter_type == "모든 근무조" or \
                       (filter_type == "오전조" and shift_name == "오전") or \
                       (filter_type == "오후조" and shift_name == "오후") or \
                       (filter_type == "야간조" and shift_name == "야간"):

                        color = "black"
                        if day.weekday() in [5, 6]:
                            color = "red"

                        on_duty_names = []
                        for name in names:
                            staff_info = next((s for s in self.staff_list if s['name'] == name), None)

                            is_additional_work_day = day.day in self.additional_work_days.get(name, [])
                            is_night_support_day = name in self.night_support_assignments and day.day in self.night_support_assignments[name]

                            if is_additional_work_day:
                                on_duty_names.append(f"{name}(근무)")
                                color = "blue"
                            elif is_night_support_day:
                                on_duty_names.append(f"{name}(지)")
                                color = "blue"
                            else:
                                on_duty_names.append(name)

                        if on_duty_names:
                            unique_on_duty_names = sorted(list(set(on_duty_names)))
                            display_text = f"{shift_name}: {', '.join(unique_on_duty_names)}"
                            ttk.Label(cell_frame, text=display_text, font=("Arial", 8), foreground=color).pack(anchor="w", padx=1, pady=0)

            col += 1
            if col > 6:
                col = 0
                row += 1

        for i in range(7):
            cal_grid.grid_columnconfigure(i, weight=1)
        for i in range(1, row):
            cal_grid.grid_rowconfigure(i, weight=1)

        legal_info_label = ttk.Label(self.calendar_frame, text="* 법정근로시간: 1일 8시간, 주 40시간 (연장근로 시 주 12시간 한도)", font=("Arial", 8, "italic"), foreground="grey")
        legal_info_label.pack(side=tk.BOTTOM, pady=5)


    def get_staff_for_day(self, day):
        staff_on_duty = {}

        # 1. 야간조 배정 (야간 지원 우선)
        night_staff_list = []
        night_support_person = None
        if self.current_mode == "기본 모드 (5인 이상)":
            for name, days in self.night_support_assignments.items():
                if day.day in days:
                    night_support_person = name
                    break

        if night_support_person:
            night_staff_list.append(night_support_person)
        else:
            if self.night_shift_index < len(self.fixed_night_schedule):
                night_person_name = self.fixed_night_schedule[self.night_shift_index]
                night_staff_list.append(night_person_name)
                self.night_shift_index += 1

        if night_staff_list:
            staff_on_duty["야간"] = sorted(list(set(night_staff_list)))

        # 2. 주간조 배정 (오전, 오후, 지원)
        for staff in self.staff_list:
            name = staff['name']
            shift_type = staff['shift']

            if shift_type.startswith('야간'):
                continue

            if night_support_person and name == night_support_person:
                continue

            is_additional_work_day = day.day in self.additional_work_days.get(name, [])
            is_additional_off_day = day.day in self.additional_off_days.get(name, [])
            is_regular_off_day = day.weekday() in staff.get('off_days', [])

            if is_additional_work_day:
                assigned_shift = '오전' if shift_type.startswith('오전') or shift_type == '지원' else '오후'
                if assigned_shift not in staff_on_duty:
                    staff_on_duty[assigned_shift] = []
                staff_on_duty[assigned_shift].append(name)
            elif not is_regular_off_day and not is_additional_off_day:
                assigned_shift = None
                # '오전B', '오후B' 근무조 추가 로직
                if shift_type == '오전A' or shift_type == '오전B':
                    assigned_shift = '오전'
                elif shift_type == '오후A' or shift_type == '오후B':
                    assigned_shift = '오후'
                # '지원' 근무조는 비상 모드(5인 이하)에서만 작동
                elif self.current_mode == "비상 모드 (5인 이하)" and shift_type == '지원':
                    assigned_shift = self.support_rotation_map.get(day.weekday())

                if assigned_shift and assigned_shift != '휴무':
                    if assigned_shift not in staff_on_duty:
                        staff_on_duty[assigned_shift] = []
                    staff_on_duty[assigned_shift].append(name)

        if "오전" in staff_on_duty:
            staff_on_duty["오전"] = sorted(staff_on_duty["오전"])
        if "오후" in staff_on_duty:
            staff_on_duty["오후"] = sorted(staff_on_duty["오후"])

        return staff_on_duty

    
    def start_new_work_auto(self, year=None, month=None):
        """
        새로운 월의 근무표 작업을 시작하는 내부 함수 (초기화 로직 포함)
        """
        # Load the latest work_stats to get the year and month
        try:
            with open(WORK_STATS_FILE, "r", encoding="utf-8") as f: # 통일된 변수 사용
                temp_work_stats = json.load(f)
            # Use the year and month from the file, if they exist.
            self.current_year = temp_work_stats.get('year', datetime.now().year)
            self.current_month = temp_work_stats.get('month', datetime.now().month)
        except (FileNotFoundError, json.JSONDecodeError):
            # If file doesn't exist or is corrupt, use the current system date as a fallback.
            now = datetime.now()
            self.current_year = now.year
            self.current_month = now.month

        # 기존 월의 야간 지원 데이터는 초기화 (새 월이 시작되면 이전 월의 야간 지원은 무효)
        self.night_support_assignments = {}
        save_night_support_data(self.night_support_assignments, self.current_year, self.current_month)

        # 휴무/근무 지정 데이터 초기화 (요청에 따라 새 월 시작 시 초기화)
        self.additional_off_days = {}
        save_additional_off_days(self.additional_off_days)

        self.additional_work_days = {}
        save_additional_work_days(self.additional_work_days)

        # 새로운 통계 데이터 초기화
        new_work_stats = {
            "year": self.current_year,
            "month": self.current_month,
            "workdays": 0,
            "total_hours": 0,
            "night_support_days": 0,
            "night_leave_hours": 0,
            "names": {s['name']: {"work_days": 0, "work_hours": 0} for s in self.staff_list}
        }
        self.work_stats = new_work_stats
        # Update work_stats_data.json with the new year and month
        save_work_stats(self.work_stats)

        self.update_calendar()
        self.update_off_days_section()
        messagebox.showinfo("시작 완료", f"{self.current_year}년 {self.current_month}월 근무표 작성을 시작합니다.")


    


if __name__ == "__main__":
    staff_list, night_support_assignments, additional_off_days, additional_work_days, work_stats, current_year, current_month = load_data()

    app = WorkScheduleApp(staff_list, night_support_assignments, additional_off_days, additional_work_days, work_stats, current_year, current_month)
    app.mainloop()