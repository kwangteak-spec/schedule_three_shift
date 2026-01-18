from flask import Flask, render_template, request, redirect, url_for, jsonify
from db_manager import DBManager
import datetime
import calendar
from holidays import is_holiday

app = Flask(__name__)
db = DBManager()

def get_monthly_target(year=None, month=None):
    now = datetime.datetime.now()
    if year is None:
        year = now.year
    if month is None:
        month = now.month
    
    # Calculate weekdays (Mon=0, Sun=6)
    num_days = calendar.monthrange(year, month)[1]
    workdays = 0
    for day in range(1, num_days + 1):
        # 0=Mon, ..., 6=Sun
        d = datetime.date(year, month, day)
        weekday = d.weekday()
        if weekday < 5 and not is_holiday(d):  # 0-4 are weekdays, and it's not a holiday
            workdays += 1
            
    total_hours = workdays * 8
    
    # User-requested calculations
    quotient = total_hours // 12
    calc_1 = quotient * 2
    # The user wants Night Shift Support Days = Total Days - (Quotient * 2)
    support_days = num_days - calc_1
    
    target = {
        'num_days': num_days,
        'workdays': workdays,
        'total_hours': total_hours,
        'calc_1': calc_1,
        'support_days': support_days
    }
    return year, month, target

@app.route('/')
def index():
    year = int(request.args.get('year', datetime.datetime.now().year))
    month = int(request.args.get('month', datetime.datetime.now().month))
    _, _, target = get_monthly_target(year, month)
    staff_list = db.fetch_staff_list()
    
    # Fetch night support assignments for the current month
    exceptions = db.fetch_schedule_exceptions(year, month)
    night_supports = [e for e in exceptions if e['event_type'] == 'NIGHT_SUPPORT']
    
    return render_template('index.html', 
                          year=year, 
                          month=month, 
                          target=target, 
                          staff_list=staff_list,
                          night_supports=night_supports)

@app.route('/delete_night_support', methods=['POST'])
def delete_night_support():
    data = request.json
    name = data['name']
    date_str = data['date'] # Expected format 'YYYY-MM-DD'
    
    db.delete_exception(name, date_str, 'NIGHT_SUPPORT')
    return jsonify({'status': 'success'})


    
def calculate_schedule_data(year, month):
    _, _, target = get_monthly_target(year, month)
    
    # Fetch Exceptions (Support, Leave, etc.)
    exceptions = db.fetch_schedule_exceptions(year, month)
    
    # Map exceptions: {day_int: {staff_name: type}}
    # Map exceptions: {day_int: {staff_name: set(types)}}
    exception_map = {}
    support_days_count = 0
    
    for exc in exceptions:
        d = exc['event_date'] 
        day_val = d.day if isinstance(d, datetime.date) else int(str(d).split('-')[-1])
            
        if day_val not in exception_map:
            exception_map[day_val] = {}
        
        name = exc['staff_name']
        if name not in exception_map[day_val]:
            exception_map[day_val][name] = set()
        exception_map[day_val][name].add(exc['event_type'])
        
        if exc['event_type'] == 'NIGHT_SUPPORT':
            support_days_count += 1
    
    stats = {
        'legal_days': target['workdays'],
        'hours': target['total_hours'],
        'support': support_days_count,
        'leave': 0
    }
    
    staff_list = db.fetch_staff_list()
    month_days = calendar.monthcalendar(year, month)
    
    # Generate schedule map
    schedule_map = {}
    
    def parse_off_days(off_str):
        if not off_str: return []
        try:
            normalized = off_str.replace('|', ',')
            return [int(d) for d in normalized.split(',') if d.strip().isdigit()]
        except: return []

    staff_by_shift = {'A': [], 'B': [], 'C': [], 'N': []}
    for staff in staff_list:
        sType = staff['shift_type']
        if sType in staff_by_shift:
            staff_by_shift[sType].append(staff)
    staff_by_shift['C'].sort(key=lambda x: x['name'])
    
    c_rotation_idx = 0
    c_worked_days = 0

    num_days = calendar.monthrange(year, month)[1]
    for day in range(1, num_days + 1):
        weekday = (calendar.weekday(year, month, day) + 1) % 7
        daily_schedule = {'A': [], 'B': [], 'C': [], 'N': []}
        day_exceptions = exception_map.get(day, {})
        
        c_staff = staff_by_shift['C']
        
        daily_schedule = {'A': [], 'B': [], 'C': [], 'N': []}
        
        for staff in staff_list:
            name = staff['name']
            excs = day_exceptions.get(name, set())
            sType = staff['shift_type']
            
            # Check for PARTIAL_OFF (reduced hours)
            if 'PARTIAL_OFF' in excs:
                # Find the PARTIAL_OFF exception to get the hours
                partial_hours = None
                for exc in exceptions:
                    if exc['staff_name'] == name and exc['event_type'] == 'PARTIAL_OFF':
                        d = exc['event_date']
                        day_val = d.day if isinstance(d, datetime.date) else int(str(d).split('-')[-1])
                        if day_val == day:
                            try:
                                partial_hours = int(exc['description']) if exc['description'] else None
                            except:
                                pass
                            break
                
                if partial_hours:
                    daily_schedule[sType].append(f"{name}({partial_hours})")
                continue
            
            if 'OFF' in excs:
                continue
            
            if 'NIGHT_SUPPORT' in excs:
                daily_schedule['C'].append(f"{name}(지)")
                continue
                
            if 'EXTRA_HOURS' in excs:
                daily_schedule[sType].append(f"{name}(추)")
                continue
            
            # Default logic for non-C shifts
            if sType != 'C':
                if weekday not in parse_off_days(staff.get('off_days', '')):
                    daily_schedule[sType].append(name)
        
        # 3. Process C shift Rotation
        manual_c_exists = any('(추)' in n or '(지)' in n for n in daily_schedule['C'])
        
        if c_staff and not manual_c_exists:
            attempts = 0
            while attempts < len(c_staff):
                worker = c_staff[c_rotation_idx]
                worker_name = worker['name']
                already_working = any(worker_name in n for shift_list in daily_schedule.values() for n in shift_list)
                
                if 'OFF' not in day_exceptions.get(worker_name, set()) and not already_working:
                    daily_schedule['C'].append(worker_name)
                    break
                
                c_rotation_idx = (c_rotation_idx + 1) % len(c_staff)
                c_worked_days = 0 
                attempts += 1
            
            c_worked_days += 1
            if c_worked_days >= 2:
                c_rotation_idx = (c_rotation_idx + 1) % len(c_staff)
                c_worked_days = 0

        schedule_map[day] = daily_schedule

    # Calculate individual work totals
    work_totals = {staff['name']: {'days': 0, 'hours': 0} for staff in staff_list}
    
    for day in schedule_map:
        for shift in ['A', 'B', 'C', 'N']:
            for entry in schedule_map[day][shift]:
                is_support = '(지)' in entry
                is_manual = '(추)' in entry
                
                import re
                partial_match = re.search(r'\((\d+)\)$', entry)
                is_partial = partial_match and not is_support and not is_manual
                
                clean_name = entry.replace('(지)', '').replace('(추)', '')
                if is_partial:
                    clean_name = re.sub(r'\(\d+\)$', '', clean_name)
                
                if clean_name in work_totals:
                    work_totals[clean_name]['days'] += 1
                    
                    if is_partial:
                        partial_hours = int(partial_match.group(1))
                        work_totals[clean_name]['hours'] += partial_hours
                        continue
                    
                    if is_manual:
                        continue
                        
                    if is_support or shift == 'C':
                        work_totals[clean_name]['hours'] += 12
                    else:
                        work_totals[clean_name]['hours'] += 8

    # Second, add manual EXTRA_HOURS from exceptions
    for exc in exceptions:
        if exc['event_type'] == 'EXTRA_HOURS':
            name = exc['staff_name']
            if name in work_totals:
                try:
                    hours = int(exc['description']) if exc['description'] else 0
                    work_totals[name]['hours'] += hours
                except:
                    pass

    return {
        'stats': stats,
        'staff_list': staff_list, 
        'month_days': month_days,
        'schedule_map': schedule_map,
        'night_supports': [exc for exc in exceptions if exc['event_type'] == 'NIGHT_SUPPORT'],
        'work_totals': work_totals
    }

@app.route('/calendar')
def calendar_view():
    year = int(request.args.get('year', datetime.datetime.now().year))
    month = int(request.args.get('month', datetime.datetime.now().month))
    
    data = calculate_schedule_data(year, month)
    
    return render_template('calendar.html', 
                          year=year, month=month, 
                          stats=data['stats'], 
                          staff_list=data['staff_list'], 
                          month_days=data['month_days'], 
                          schedule_map=data['schedule_map'],
                          night_supports=data['night_supports'],
                          work_totals=data['work_totals'])

@app.route('/add_staff', methods=['POST'])
def add_staff():
    name = request.form['name']
    shift = request.form['shift']
    off_days = request.form['off_days']
    
    db.add_staff(name, shift, off_days)
    return redirect(url_for('index'))

@app.route('/update_staff', methods=['POST'])
def update_staff():
    original_name = request.form['original_name']
    name = request.form['name']
    shift = request.form['shift']
    off_days = request.form['off_days']
    
    db.update_staff(original_name, name, shift, off_days)
    return redirect(url_for('index'))

@app.route('/delete_staff', methods=['POST'])
def delete_staff():
    name = request.form['name']
    db.delete_staff(name)
    return redirect(url_for('index'))

@app.route('/save_extra_hours', methods=['POST'])
def save_extra_hours():
    try:
        data = request.json
        name = data.get('name')
        day = data.get('day')
        hours = data.get('hours')
        year = data.get('year')
        month = data.get('month')
        
        if not all([name, day, hours, year, month]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
            
        date_str = f"{year}-{month:02d}-{int(day):02d}"
        # When adding work, clear 'OFF' status
        db.update_exception(name, date_str, 'EXTRA_HOURS', value=hours, clear_types=['OFF'])
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"Error in save_extra_hours: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/delete_extra_hours', methods=['POST'])
def delete_extra_hours():
    data = request.json
    name = data['name']
    day = data['day']
    year = data['year']
    month = data['month']
    
    date_str = f"{year}-{month:02d}-{int(day):02d}"
    db.delete_exception(name, date_str, 'EXTRA_HOURS')
    return jsonify({'status': 'success'})

@app.route('/save_night_support', methods=['POST'])
def save_night_support():
    data = request.json
    name = data['name']
    year = data['year']
    month = data['month']
    day = data['day']
    
    date_str = f"{year}-{month:02d}-{int(day):02d}"
    db.update_exception(name, date_str, 'NIGHT_SUPPORT')
    return jsonify({'status': 'success'})

@app.route('/exclude_from_work', methods=['POST'])
def exclude_from_work():
    data = request.json
    name = data['name']
    day = data['day']
    year = data['year']
    month = data['month']
    
    # Get staff info to determine shift type
    staff_list = db.fetch_staff_list()
    staff = next((s for s in staff_list if s['name'] == name), None)
    
    if not staff:
        return jsonify({'status': 'error', 'message': 'Staff not found'}), 404
    
    # Calculate reduced hours based on shift type
    # C shift = 12h normally, so reduced = 4h
    # A, B, N shift = 8h normally, so reduced = 4h
    shift_type = staff['shift_type']
    if shift_type == 'C':
        reduced_hours = 4  # 12h - 8h = 4h remaining
    else:
        reduced_hours = 4  # 8h - 4h = 4h remaining
    
    date_str = f"{year}-{month:02d}-{int(day):02d}"
    # When excluding work, use PARTIAL_OFF with reduced hours
    db.update_exception(name, date_str, 'PARTIAL_OFF', value=reduced_hours, clear_types=['EXTRA_HOURS', 'NIGHT_SUPPORT', 'OFF'])
    return jsonify({'status': 'success'})

@app.route('/delete_exception', methods=['POST'])
def delete_exception():
    data = request.json
    name = data['name']
    day = data['day']
    year = data['year']
    month = data['month']
    exc_type = data.get('type')
    
    date_str = f"{year}-{month:02d}-{int(day):02d}"
    db.delete_exception(name, date_str, exc_type)
    return jsonify({'status': 'success'})

@app.route('/export_excel')
def export_excel():
    year = int(request.args.get('year', datetime.datetime.now().year))
    month = int(request.args.get('month', datetime.datetime.now().month))
    
    data = calculate_schedule_data(year, month)
    schedule_map = data['schedule_map']
    
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 {month}월 근무표"
    
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    
    # Title
    ws['A1'] = "CU편의점"
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(size=14, bold=True)
    
    ws['G1'] = f"{year}년"
    ws['H1'] = f"{month}월"
    ws['H1'].alignment = Alignment(horizontal='right')
    
    # Headers
    headers = ["일", "월", "화", "수", "목", "금", "토"]
    week_calendar = calendar.monthcalendar(year, month)
    
    current_row = 2
    
    # Set Column Widths
    ws.column_dimensions['A'].width = 15 # Label column
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
        
    # Iterate through weeks
    for week_idx, week in enumerate(week_calendar):
        # 1. Header Row (Days of Week) - Only for the first week? 
        # The image shows headers '일', '월'... above the first week block.
        # But actually the image shows structure:
        # [Header Row: 일...토]
        # [Data Block: Date, Morning, Afternoon, Night]
        # And repeats Date row for next week.
        
        # Let's follow the image: "일 월 화 ... 토" header is usually at the very top of the grid.
        if week_idx == 0:
            ws.cell(row=current_row, column=1, value="") # Corner
            for i, day_name in enumerate(headers):
                c = ws.cell(row=current_row, column=i+2, value=day_name)
                c.alignment = center_align
                c.border = thin_border
                c.font = bold_font
            current_row += 1
            
        # Rows: Date, Morning(오전근무), Afternoon(오후근무), Night(야간근무)
        row_labels = ["일자", "오전근무", "오후근무", "야간근무"]
        shift_keys = [None, 'A', 'B', 'C'] # None for Date row
        
        start_row = current_row
        
        for r_offset, label in enumerate(row_labels):
            # Label Cell
            label_cell = ws.cell(row=start_row + r_offset, column=1, value=label)
            label_cell.alignment = center_align
            label_cell.border = thin_border
            label_cell.font = bold_font
            
            # Data Cells
            for day_idx, day in enumerate(week):
                col_idx = day_idx + 2
                cell = ws.cell(row=start_row + r_offset, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align
                
                if day == 0:
                    continue
                
                if r_offset == 0: # Date Row
                    cell.value = day
                    cell.font = bold_font
                else: # Shift Rows
                    shift_key = shift_keys[r_offset]
                    if day in schedule_map:
                        workers = schedule_map[day][shift_key]
                        # Clean up names for display (remove (추), (지) etc if needed, or keep them)
                        # Image shows clean names. Let's keep markers as they are useful info.
                        cell.value = "\n".join(workers)
        
        current_row += 4 # Move to next block

    # Finalize
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, download_name=f"Schedule_{year}_{month}.xlsx", as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
